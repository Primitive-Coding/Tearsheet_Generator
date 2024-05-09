# Operating system
import os

# Number manipulation
import pandas as pd
import numpy as np

# Date & Time
import datetime as dt

# Stock libraries.
import yfinance as yf
from Tearsheet.forecast_model import ForecastModel

# Excel libaries
from Tearsheet.tearsheet_configs import config_1
from Tearsheet.tearsheet_utils import col_map
from openpyxl import Workbook, load_workbook

from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis

cwd = os.getcwd()

# Scraper imports
from Scraper.alpha_vantage import AlphaVantage

annual_params = ["A", "a", "Annual", "annual"]
quarterly_params = ["Q", "q", "Quarterly", "quarterly", "Quarter", "quarter"]


class Tearsheet:
    def __init__(self, ticker: str, freq: str = "A") -> None:
        self.ticker = ticker.upper()
        if freq in annual_params:
            self.freq = "A"
            self.folder = "Annual"
        elif freq in quarterly_params:
            self.freq = "Q"
            self.folder = "Quarter"

        self.path_to_data = f"D:\\STOCK_ANALYSIS_DATASET\\{self.ticker}\\{self.folder}"
        self.forecast_model = ForecastModel(self.ticker, self.freq, predict_ratio=True)

        # Define fonts and fills
        self.header_font = Font(bold=True, size=16, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        self.subheader_font = Font(bold=True, size=12)
        self.subheader_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        self.data_font = Font(size=11)
        self.bold_font = Font(size=11, bold=True)
        self.data_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )

        # Define borders
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        self.top_thin_border = Border(top=Side(style="thin"))
        self.right_thin_border = Border(right=Side(style="thin"))
        self.bottom_thin_border = Border(bottom=Side(style="thin"))
        self.left_thin_border = Border(left=Side(style="thin"))
        self.top_left_corner_thin_border = Border(
            top=Side(style="thin"), left=Side(style="thin")
        )
        self.bottom_left_corner_thin_border = Border(
            bottom=Side(style="thin"), left=Side(style="thin")
        )
        self.top_right_corner_thin_border = Border(
            top=Side(style="thin"), right=Side(style="thin")
        )
        self.bottom_right_corner_thin_border = Border(
            bottom=Side(style="thin"), right=Side(style="thin")
        )
        # Thick borders
        self.top_thick_border = Border(top=Side(style="thick"))
        self.right_thick_border = Border(right=Side(style="thick"))
        self.bottom_thick_border = Border(bottom=Side(style="thick"))
        self.left_thick_border = Border(left=Side(style="thick"))
        self.top_left_corner_thick_border = Border(
            top=Side(style="thick"), left=Side(style="thick")
        )
        self.bottom_left_corner_thick_border = Border(
            bottom=Side(style="thick"), left=Side(style="thick")
        )
        self.top_right_corner_thick_border = Border(
            top=Side(style="thick"), right=Side(style="thick")
        )
        self.bottom_right_corner_thick_border = Border(
            bottom=Side(style="thick"), right=Side(style="thick")
        )

        self.basic_format = "{:,.0f}"
        self.decimal_format = "{:,.2f}"
        self.dollar_decimal_format = "${:,.2f}"
        self.dollar_basic_format = "${:,.0f}"
        self.pct_format = "{:,.2f}%"
        self.pct_basic_format = "{:,.0f}%"

        # ALignments
        self.left_align = Alignment(horizontal="left", vertical="center")
        self.right_align = Alignment(horizontal="right", vertical="center")
        self.center_align = Alignment(horizontal="center", vertical="center")
        # Predefined alignment sets
        # Column custom widths
        self.main_col_width = 40
        self.sub_col_width = 20

        # Define table configuration
        self.table_config = config_1

    def set_all_statements(self) -> None:
        """
        Sets every financial statement as a class method.
        """
        self.income_statement = self.stock_analysis_scraper.get_income_statement()
        self.balance_sheet = self.stock_analysis_scraper.get_balance_sheet()
        self.cash_flow = self.stock_analysis_scraper.get_cash_flow()
        self.ratios = self.stock_analysis_scraper.get_ratios()

    def get_returns(self, df: pd.DataFrame):
        """
        Get stock returns between fiscal period.

        dates: list of fiscal dates.

        """
        # Reverse dates so old dates are on the left, and new dates are on the right.

        dates = df.columns

        if dates[0] > dates[1]:
            dates = dates[::-1]

        candles = self.get_candles()
        returns = []
        stock_high = []
        stock_low = []
        stock_average = []
        index = 0
        for d in dates:
            if index == 0:
                returns.append(np.nan)
                stock_high.append(np.nan)
                stock_low.append(np.nan)
                stock_average.append(np.nan)
            else:
                start_date = dates[index - 1]
                end_date = d
                price_data = candles.loc[dates[index - 1] : d]
                # Extract stock prices
                stock_high.append(round(price_data["Adj Close"].max(), 2))
                stock_low.append(round(price_data["Adj Close"].min(), 2))
                stock_average.append(round(price_data["Adj Close"].mean(), 2))
                # Calculate returns for the period
                try:
                    start_price = price_data["Adj Close"].iloc[0]
                    end_price = price_data["Adj Close"].iloc[-1]
                    r = ((end_price - start_price) / abs(start_price)) * 100
                except IndexError:
                    r = np.nan
                returns.append(r)
            index += 1

        return {
            "returns": returns,
            "high": stock_high,
            "low": stock_low,
            "average": stock_average,
        }

    def create_tearsheet(self):

        self.stock_info = yf.Ticker(self.ticker).get_info()
        self.historical_data = yf.download(self.ticker)
        self.forecast_model.set_all_statements()

        # self.set_all_statements()
        path = f"Tearsheet\\Tearsheets\\{self.ticker}_tearsheet.xlsx"
        try:
            self.wb = load_workbook(path)
        except FileNotFoundError:
            self.wb = Workbook()

        sheet_names = self.wb.sheetnames
        cur_date = dt.datetime.now().date().strftime("%Y-%m-%d")

        if cur_date not in sheet_names:
            self.ws = self.wb.create_sheet(title=f"{cur_date}")
        else:
            self.ws = self.wb.active

        # Create header for tearsheet.
        self.create_header()
        # Create market profile section.
        self.create_market_profile()
        # Create key financials section.
        self.create_key_financials()
        # Create valuation method section.
        self.create_valuation_method()
        # Create summary section.
        self.create_business_summary()
        # Create investment highligh section.
        self.create_investment_highlights()
        # Create historical growth section.
        self.create_historical_growth()
        # Create historical ratio section.
        self.create_historical_ratios()
        # Create rebuilt share section.
        self.create_rebuilt_share()

        self.wb.save(path)

    def create_header(self):
        # Example: Applying header style
        for col in range(1, 10):  # Adjust the range based on your needs
            cell = self.ws[f"{get_column_letter(col)}1"]
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Set column widths
            if col == 1:
                self.ws.column_dimensions[get_column_letter(col)].width = (
                    self.main_col_width
                )
            else:
                self.ws.column_dimensions[get_column_letter(col)].width = (
                    self.sub_col_width
                )

        self.ws["A1"] = self.stock_info["longName"]

    def create_market_profile(self):
        start_index = self.table_config["market_profile"]["row"]
        one_year_ago = self.get_x_year_ago(1)

        """--- Instructions for table insertion ---"""
        headers = ["Market Profile", ""]

        try:
            dividend_yield = self.pct_format.format(
                (
                    abs(self.forecast_model.cash_flow.loc["Dividends Paid"].iloc[-1])
                    / self.forecast_model.ratios.loc["Market Capitalization"].iloc[-1]
                )
                * 100
            )
        except KeyError:
            dividend_yield = self.pct_format.format(0)

        instructions = {
            "header": {
                "values": headers,
            },
            "body": {
                "close": {
                    "values": [
                        "Closing Price",
                        self.dollar_decimal_format.format(
                            self.historical_data["Adj Close"].iloc[-1]
                        ),
                    ],
                },
                "year_high": {
                    "values": [
                        "52-Week High",
                        self.dollar_decimal_format.format(
                            self.historical_data.loc[one_year_ago:]["Adj Close"].max()
                        ),
                    ],
                },
                "year_low": {
                    "values": [
                        "52-Week Low",
                        self.dollar_decimal_format.format(
                            self.historical_data.loc[one_year_ago:]["Adj Close"].min()
                        ),
                    ],
                },
                "shares": {
                    "values": [
                        "Shares Outstanding (Mil)",
                        self.basic_format.format(
                            self.forecast_model.income_statement.loc[
                                "Shares Outstanding (Basic)"
                            ].iloc[-1]
                        ),
                    ],
                },
                "eps": {
                    "values": [
                        "EPS (TTM)",
                        self.dollar_decimal_format.format(
                            self.forecast_model.income_statement.loc[
                                "EPS (Basic)"
                            ].iloc[-1]
                        ),
                    ],
                },
                "marketcap": {
                    "values": [
                        "Marketcap (Mil)",
                        self.dollar_basic_format.format(
                            self.forecast_model.ratios.loc[
                                "Market Capitalization"
                            ].iloc[-1]
                        ),
                    ],
                },
                "avg_volume": {
                    "values": [
                        "Average Volume (TTM)",
                        self.basic_format.format(
                            self.historical_data.loc[one_year_ago:]["Volume"].max()
                        ),
                    ],
                },
                "beta": {
                    "values": ["Beta", "NaN"],
                },
                "financial_health": {
                    "values": ["Financial Health", "NaN"],
                },
                "country": {
                    "values": ["Country", self.stock_info["country"]],
                },
                "p/e": {
                    "values": [
                        "P/E",
                        self.decimal_format.format(
                            self.forecast_model.ratios.loc["PE Ratio"].iloc[-1]
                        ),
                    ],
                },
                "gross_margin": {
                    "values": [
                        "Gross Margin",
                        self.pct_format.format(
                            (
                                self.forecast_model.income_statement.loc[
                                    "Gross Margin"
                                ].iloc[-1]
                            )
                            * 100
                        ),
                    ],
                },
                "operating_margin": {
                    "values": [
                        "Operating Margin",
                        self.pct_format.format(
                            (
                                self.forecast_model.income_statement.loc[
                                    "Operating Margin"
                                ].iloc[-1]
                            )
                            * 100
                        ),
                    ],
                },
                "net_margin": {
                    "values": [
                        "Net Margin",
                        self.pct_format.format(
                            (
                                self.forecast_model.income_statement.loc[
                                    "Profit Margin"
                                ].iloc[-1]
                            )
                            * 100
                        ),
                    ],
                },
                "fcf_margin": {
                    "values": [
                        "Free Cash Flow Margin",
                        self.pct_format.format(
                            (
                                self.forecast_model.income_statement.loc[
                                    "Free Cash Flow Margin"
                                ].iloc[-1]
                            )
                            * 100
                        ),
                    ],
                },
                "div_yield": {
                    "values": ["Dividend Yield", dividend_yield],
                },
            },
        }
        self.insert_table(
            start_index=start_index,
            cols=self.table_config["market_profile"]["cols"],
            instructions=instructions,
        )

    def create_key_financials(self):
        """
        Headers: A19, B19, C19
        """
        start_index = self.table_config["key_financials"]["row"]
        future_index = 0

        """--- Get information for table ---"""
        revenue = self.forecast_model.income_statement.loc["Revenue"]
        proj_revenue = self.forecast_model.create_forecast3(revenue)["forecast"].iloc[
            future_index
        ]
        op_income = self.forecast_model.income_statement.loc["Operating Income"]
        proj_op_income = self.forecast_model.create_forecast3(op_income)[
            "forecast"
        ].iloc[future_index]
        net_income = self.forecast_model.income_statement.loc["Net Income"]
        proj_net_income = self.forecast_model.create_forecast3(net_income)[
            "forecast"
        ].iloc[future_index]
        eps = self.forecast_model.income_statement.loc["EPS (Basic)"]
        proj_eps = self.forecast_model.create_forecast3(eps)["forecast"].iloc[
            future_index
        ]
        fcf = self.forecast_model.cash_flow.loc["Free Cash Flow"]
        proj_fcf = self.forecast_model.create_forecast3(fcf)["forecast"].iloc[
            future_index
        ]
        try:
            dividends = self.forecast_model.income_statement.loc["Dividend Per Share"]
            proj_dividends = self.forecast_model.create_forecast3(dividends)[
                "forecast"
            ].iloc[future_index]
            dividends = dividends.iloc[-1]
        except KeyError:
            dividends = 0
            proj_dividends = 0
        """--- Instructions for table insertion ---"""
        headers = ["Key Financials", "TTM", "Projected (1-Year)"]
        instructions = {
            "header": {
                "values": headers,
            },
            "body": {
                "revenue": {
                    "values": [
                        "Revenue (Mil)",
                        self.dollar_basic_format.format(revenue.iloc[-1]),
                        self.dollar_basic_format.format(proj_revenue),
                    ],
                },
                "operating_income": {
                    "values": [
                        "Operating Income (Mil)",
                        self.dollar_basic_format.format(op_income.iloc[-1]),
                        self.dollar_basic_format.format(proj_op_income),
                    ],
                },
                "net_income": {
                    "values": [
                        "Net Income (Mil)",
                        self.dollar_basic_format.format(net_income.iloc[-1]),
                        self.dollar_basic_format.format(proj_net_income),
                    ],
                },
                "EPS": {
                    "values": [
                        "EPS",
                        self.dollar_decimal_format.format(eps.iloc[-1]),
                        self.dollar_decimal_format.format(proj_eps),
                    ],
                },
                "fcf": {
                    "values": [
                        "Free Cash Flow (Mil)",
                        self.dollar_basic_format.format(fcf.iloc[-1]),
                        self.dollar_basic_format.format(proj_fcf),
                    ],
                },
                "dividends": {
                    "values": [
                        "Dividends Per Share",
                        self.dollar_decimal_format.format(dividends),
                        self.dollar_decimal_format.format(proj_dividends),
                    ],
                },
            },
        }
        """--- Instructions for table insertion ---"""
        self.insert_table(
            start_index=start_index,
            cols=self.table_config["key_financials"]["cols"],
            instructions=instructions,
        )

    def create_valuation_method(self):
        """ """
        start_index = self.table_config["valuation_method"]["row"]
        future_index = -1  # Index to use for forecasting value.

        """--- Get information for table ---"""
        headers = ["Valuation Method", "Weight", "Projected (1-Year)"]
        revenue = self.forecast_model.get_revenue_valuation()[
            "projected_share_price"
        ].iloc[future_index]

        income = self.forecast_model.get_earnings_valuation()[
            "projected_share_price"
        ].iloc[future_index]

        equity = self.forecast_model.get_equity_valuation()[
            "projected_share_price"
        ].iloc[future_index]

        fcf = self.forecast_model.get_fcf_valuation()["projected_share_price"].iloc[
            future_index
        ]

        """--- Instructions for table insertion ---"""
        instructions = {
            "header": {
                "values": headers,
            },
            "body": {
                "revenue": {
                    "values": [
                        "Revenue",
                        "",
                        self.dollar_decimal_format.format(revenue),
                    ],
                },
                "income": {
                    "values": [
                        "Net Income",
                        "",
                        self.dollar_decimal_format.format(income),
                    ],
                },
                "equity": {
                    "values": ["Equity", "", self.dollar_decimal_format.format(equity)],
                },
                "fcf": {
                    "values": [
                        "Free Cash Flow",
                        "",
                        self.dollar_decimal_format.format(fcf),
                    ],
                },
            },
        }

        k = list(instructions["body"].keys())
        total = len(k)
        average_weight = (1 / total) * 100
        total_weight = average_weight * total

        for i in instructions["body"].keys():
            instructions["body"][i]["values"][1] = self.pct_basic_format.format(
                average_weight
            )

        # Calculate share average share predictions.
        share_sum = 0
        for i in instructions["body"].keys():
            val = instructions["body"][i]["values"][-1]
            val = float(val.replace("$", "").replace(",", ""))
            share_sum += val

        # Insert data into table
        self.insert_table(
            start_index=start_index,
            cols=self.table_config["valuation_method"]["cols"],
            instructions=instructions,
        )

        # Implement implied value data.
        end_index = start_index + len(instructions["body"].keys()) + 1
        self.ws[
            f"{self.table_config['valuation_method']['cols'][0]}{end_index}"
        ].value = "Implied Value: "
        self.ws[
            f"{self.table_config['valuation_method']['cols'][1]}{end_index}"
        ].value = self.pct_basic_format.format(total_weight)
        self.ws[
            f"{self.table_config['valuation_method']['cols'][2]}{end_index}"
        ].value = self.dollar_decimal_format.format(share_sum / len(k))
        self.ws[
            f"{self.table_config['valuation_method']['cols'][0]}{end_index}"
        ].alignment = self.left_align
        self.ws[
            f"{self.table_config['valuation_method']['cols'][1]}{end_index}"
        ].alignment = self.right_align
        self.ws[
            f"{self.table_config['valuation_method']['cols'][2]}{end_index}"
        ].alignment = self.right_align
        self.ws[
            f"{self.table_config['valuation_method']['cols'][0]}{end_index}"
        ].border = self.top_thin_border
        self.ws[
            f"{self.table_config['valuation_method']['cols'][1]}{end_index}"
        ].border = self.top_thin_border
        self.ws[
            f"{self.table_config['valuation_method']['cols'][2]}{end_index}"
        ].border = self.top_thin_border

    def create_business_summary(self):
        cols1 = ["C", "D"]
        cols2 = ["E", "F"]
        start_index = 2
        headers1 = ["Business Summary", ""]
        instructions1 = {
            "header": {
                "values": headers1,
            },
            "body": {
                "exchange": {
                    "values": ["NYSE", self.ticker],
                },
                "industry": {
                    "values": ["Industry", self.stock_info["industry"]],
                },
                "Sector": {
                    "values": ["Sector", self.stock_info["sector"]],
                },
            },
        }
        instructions2 = {
            "header": {
                "values": [""] * len(headers1),
            },
            "body": {
                "growth_rating": {
                    "values": ["Growth Rating", "NaN"],
                },
                "dividend_rating": {
                    "values": ["Dividend Rating", "NaN"],
                },
                "five_year_growth_price": {"values": ["Five Year Growth Price", "NaN"]},
            },
        }

        # Insert instructions 1
        self.insert_table(
            start_index=start_index, cols=cols1, instructions=instructions1
        )
        # Insert instructions 2
        self.insert_table(
            start_index=start_index, cols=cols2, instructions=instructions2
        )

    def create_investment_highlights(self):
        config_key = "investment_highlights"
        start_index = self.table_config[config_key]["row"]
        cols = self.table_config[config_key]["cols"]

        self.fill_col_borders(
            col=cols[0],
            start_index=start_index,
            border=self.left_thick_border,
            r=(6, 18),
        )

        cells = [f"{c}{start_index}" for c in cols]
        self.insert_col_range(
            cells=cells,
            values=["Investment Highlights", "", "", ""],
            fonts=[self.subheader_font] * len(cells),
            fill=self.subheader_fill,
            borders=[None] * len(cells),
            alignments=[
                self.center_align,
            ]
            * len(cells),
        )
        end_row = start_index + 12
        # start_column=3 // C
        # end_column=6 // F
        self.ws.merge_cells(
            start_row=start_index + 1,
            start_column=col_map[cols[0]],
            end_row=end_row,
            end_column=col_map[cols[-1]],
        )
        top_left_cell = self.ws[f"C{start_index + 1}"]
        # Write paragraph to cell
        top_left_cell.value = self.stock_info["longBusinessSummary"]
        top_left_cell.alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="left"
        )

    def create_historical_growth(self) -> None:
        future_index = 0
        config_key = "historical_growth"
        start_index = self.table_config[config_key]["row"]
        cols = self.table_config[config_key]["cols"] {cols} Row: {start_index}")
        headers = ["Growth History", "Historical Growth", "Future Growth"]

        """--- Get Information for table ---"""
        rev_growth = self.forecast_model.income_statement.loc["Revenue Growth (YoY)"]
        average_rev_growth = self.pct_format.format(rev_growth.mean() * 100)
        future_rev_growth = self.pct_format.format(
            (
                self.forecast_model.create_forecast3(rev_growth)["forecast"].iloc[
                    future_index
                ]
            )
            * 100
        )
        # Calculate operating income forecasts.
        operating_income = self.forecast_model.income_statement.loc[
            "Operating Income"
        ].pct_change()
        average_operating_growth = self.pct_format.format(operating_income.mean() * 100)
        future_operating_growth = self.pct_format.format(
            (
                self.forecast_model.create_forecast3(operating_income)["forecast"].iloc[
                    future_index
                ]
            )
            * 100
        )
        # Calculate net income forecasts.
        net_income = self.forecast_model.income_statement.loc["Net Income"].pct_change()
        average_net_inc_growth = self.pct_format.format(net_income.mean() * 100)
        future_net_inc_growth = self.pct_format.format(
            (
                self.forecast_model.create_forecast3(net_income)["forecast"].iloc[
                    future_index
                ]
            )
            * 100
        )
        # Calculate eps forecasts.
        eps = self.forecast_model.income_statement.loc["EPS (Basic)"].pct_change()
        average_eps_growth = self.pct_format.format(eps.mean() * 100)
        future_eps_growth = self.pct_format.format(
            (self.forecast_model.create_forecast3(eps)["forecast"].iloc[future_index])
            * 100
        )
        # Calculate dividends forecast.
        try:
            dividends = self.forecast_model.income_statement.loc["Dividend Growth"]
            average_div_growth = self.pct_format.format(dividends.mean() * 100)
            future_div_growth = self.pct_format.format(
                (
                    self.forecast_model.create_forecast3(dividends)["forecast"].iloc[
                        future_index
                    ]
                )
                * 100
            )
        except KeyError:
            average_div_growth = self.pct_format.format(0)
            future_div_growth = self.pct_format.format(0)
        # Calculate free cash flow forecast.
        fcf = self.forecast_model.cash_flow.loc["Free Cash Flow"].pct_change()
        average_fcf_growth = self.pct_format.format(fcf.mean() * 100)
        future_fcf_growth = self.pct_format.format(
            (self.forecast_model.create_forecast3(fcf)["forecast"].iloc[future_index])
            * 100
        )
        """--- Instructions to insert table ---"""
        instructions = {
            "header": {
                "values": headers,
            },
            "body": {
                "revenue": {
                    "values": ["Revenue", average_rev_growth, future_rev_growth],
                },
                "operating_income": {
                    "values": [
                        "Operating Income",
                        average_operating_growth,
                        future_operating_growth,
                    ],
                },
                "net_income": {
                    "values": [
                        "Net Income",
                        average_net_inc_growth,
                        future_net_inc_growth,
                    ],
                },
                "eps": {
                    "values": ["EPS", average_eps_growth, future_eps_growth],
                },
                "fcf": {
                    "values": ["Free Cash Flow", average_fcf_growth, future_fcf_growth],
                },
                "dividends": {
                    "values": ["Dividends", average_div_growth, future_div_growth],
                },
            },
        }
        """--- Insert table with instructions ---"""
        self.insert_table(start_index=start_index, cols=cols, instructions=instructions)

    def create_historical_ratios(self):
        start_index = 26
        """--- Get information for table ---"""
        ps_ratio = self.forecast_model.ratios.loc["PS Ratio"]
        pe_ratio = self.forecast_model.ratios.loc["PE Ratio"]
        pb_ratio = self.forecast_model.ratios.loc["PB Ratio"]
        pfcf_ratio = self.forecast_model.ratios.loc["P/FCF Ratio"]
        ev_fcf_ratio = self.forecast_model.ratios.loc["EV/FCF Ratio"]
        """--- Instructions to insert table ---"""
        headers = ["Historical Ratios", "Average", "High", "Low"]
        instructions = {
            "header": {
                "values": headers,
            },
            "body": {
                "p/s": {
                    "values": [
                        "Price-to-Sales",
                        self.decimal_format.format(ps_ratio.mean()),
                        self.decimal_format.format(ps_ratio.max()),
                        self.decimal_format.format(ps_ratio.min()),
                    ],
                },
                "p/e": {
                    "values": [
                        "Price-to-Earnings",
                        self.decimal_format.format(pe_ratio.mean()),
                        self.decimal_format.format(pe_ratio.max()),
                        self.decimal_format.format(pe_ratio.min()),
                    ],
                },
                "p/b": {
                    "values": [
                        "Price-to-Book",
                        self.decimal_format.format(pb_ratio.mean()),
                        self.decimal_format.format(pb_ratio.max()),
                        self.decimal_format.format(pb_ratio.min()),
                    ]
                },
                "p/fcf": {
                    "values": [
                        "Price-to-FCF",
                        self.decimal_format.format(pfcf_ratio.mean()),
                        self.decimal_format.format(pfcf_ratio.max()),
                        self.decimal_format.format(pfcf_ratio.min()),
                    ]
                },
                "ev/fcf": {
                    "values": [
                        "EV/FCF",
                        self.decimal_format.format(ev_fcf_ratio.mean()),
                        self.decimal_format.format(ev_fcf_ratio.max()),
                        self.decimal_format.format(ev_fcf_ratio.min()),
                    ]
                },
            },
        }
        # Insert table with instructions.
        self.insert_table(
            start_index=start_index,
            cols=["D", "E", "F", "G"],
            instructions=instructions,
        )

    def create_rebuilt_share(self):
        config_key = "rebuilt_share"
        start_index = self.table_config[config_key]["row"]
        cols = self.table_config[config_key]["cols"]
        """--- Information for the table ---"""
        try:
            div_per_share = self.forecast_model.income_statement.loc[
                "Dividend Per Share"
            ]
        except KeyError:
            div_per_share = 0
        book_per_share = self.forecast_model.balance_sheet.loc["Book Value Per Share"]
        eps_per_share = self.forecast_model.income_statement.loc["EPS (Basic)"]
        fcf_per_share = self.forecast_model.income_statement.loc[
            "Free Cash Flow Per Share"
        ]
        net_cash_per_share = self.forecast_model.balance_sheet.loc["Net Cash Per Share"]
        rev_per_share = (
            self.forecast_model.income_statement.loc["Revenue"]
            / self.forecast_model.income_statement.loc["Shares Outstanding (Basic)"]
        )

        rebuild_metrics = [
            book_per_share,
            eps_per_share,
            div_per_share,
            fcf_per_share,
            net_cash_per_share,
            rev_per_share,
        ]
        rebuilt_share = 0
        for rm in rebuild_metrics:
            rebuilt_share += rm

        weights = []
        total_weight = 0
        for rm in rebuild_metrics:
            weight = rm / rebuilt_share
            total_weight += weight.iloc[-1]
            weights.append(weight)

        try:
            eps_growth = self.forecast_model.income_statement.loc["EPS Growth"].mean()
        except KeyError:
            eps = self.forecast_model.income_statement.loc["EPS (Basic)"]
            index = 0
            pct_changes = 0
            for i in eps:
                if index > 0:
                    change = (i - eps[index - 1]) / abs(eps[index - 1])
                    pct_changes += change
                index += 1
            eps_growth = pct_changes / len(eps) - 1

        try:
            fcf_growth = self.forecast_model.cash_flow.loc[
                "Free Cash Flow Growth"
            ].mean()
        except KeyError:
            fcf = self.forecast_model.cash_flow.loc["Free Cash Flow"]
            index = 0
            pct_changes = 0
            for i in fcf:
                if index > 0:
                    change = (i - fcf[index - 1]) / abs(fcf[index - 1])
                    pct_changes += change
                index += 1
            fcf_growth = pct_changes / len(fcf) - 1

        rebuilt_growth = (
            self.forecast_model.ratios.loc["Return on Equity (ROE)"].mean()
            + self.forecast_model.ratios.loc["Return on Assets (ROA)"].mean()
            + self.forecast_model.ratios.loc["Return on Capital (ROIC)"].mean()
            + eps_growth
            + self.forecast_model.income_statement.loc["Revenue Growth (YoY)"].mean()
            + fcf_growth
            + self.forecast_model.ratios.loc["Buyback Yield / Dilution"].mean()
            + self.forecast_model.balance_sheet.loc["Net Cash / Debt Growth"].mean()
            + self.forecast_model.ratios.loc["Earnings Yield"].mean()
            + self.forecast_model.ratios.loc["FCF Yield"].mean()
            + (
                (
                    self.forecast_model.income_statement.loc["Gross Margin"].mean()
                    + self.forecast_model.income_statement.loc[
                        "Operating Margin"
                    ].mean()
                    + self.forecast_model.income_statement.loc["Profit Margin"].mean()
                    + self.forecast_model.income_statement.loc[
                        "Free Cash Flow Margin"
                    ].mean()
                )
                / 3
            )
        )
        # Calculate final price base on share and growth rebuilding.
        r = rebuilt_share * rebuilt_growth
        f = r + rebuilt_share
        """--- Instructions for the table ---"""
        headers = ["Rebuilt Share Price", "Per Share", "Weight"]
        instructions = {
            "header": {
                "values": headers,
            },
            "body": {
                "book/share": {
                    "values": ["Book Per Share", 0, 0],
                },
                "eps/share": {
                    "values": ["Earnings Per Share", 0, 0],
                },
                "div/share": {
                    "values": ["Dividend Per Share", 0, 0],
                },
                "fcf/share": {
                    "values": ["FCF Per Share", 0, 0],
                },
                "net_cash/share": {
                    "values": ["Net Cash/Share", 0, 0],
                },
                "rev/share": {
                    "values": ["Revenue/Share", 0, 0],
                },
                "growth_ratio": {
                    "values": ["Growth Ratio", 0, ""],
                },
                "rebuilt_value": {
                    "values": ["Rebuilt Share Price", 0, 0],
                },
            },
        }
        keys = list(instructions["body"].keys())
        index = start_index
        # Insert new values into instructions.
        for k in range(len(rebuild_metrics)):
            try:
                instructions["body"][keys[k]]["values"][1] = (
                    self.dollar_decimal_format.format(rebuild_metrics[k].iloc[-1])
                )
            except AttributeError:
                instructions["body"][keys[k]]["values"][1] = (
                    self.dollar_decimal_format.format(0)
                )
            instructions["body"][keys[k]]["values"][2] = self.pct_basic_format.format(
                weights[k].iloc[-1] * 100
            )
            index += 1
        instructions["body"]["growth_ratio"]["values"][
            1
        ] = f"{self.decimal_format.format(rebuilt_growth)}x"
        instructions["body"]["rebuilt_value"]["values"][1] = (
            self.dollar_decimal_format.format(f.iloc[-1])
        )
        instructions["body"]["rebuilt_value"]["values"][2] = (
            self.pct_basic_format.format(total_weight * 100)
        )
        """--- Insert table using instructions ---"""
        self.insert_table(
            start_index=start_index,
            cols=cols,
            instructions=instructions,
        )
        self.fill_row_borders(
            index=start_index + 7, cols=cols, border=self.top_thin_border
        )

        self.fill_row_borders(
            index=start_index + 8, cols=cols, border=self.top_thin_border
        )

    def create_stock_chart(self):
        one_year_ago = self.get_x_year_ago(1)
        chart_data = self.historical_data.loc[one_year_ago:]
        price_data = self.historical_data.loc[one_year_ago:]
        close = price_data["Adj Close"]

        chart = LineChart()
        chart.title = "52-Week Price History"
        chart.style = 13  # Predefined Excel chart style (optional)
        chart.y_axis.title = "Price per Share"
        chart.x_axis.title = "Date"
        chart.x_axis.number_format = "d-mmm"
        chart.x_axis.majorTimeUnit = "days"
        chart.x_axis = DateAxis(crossAx=100)
        chart.x_axis.title = "Date"

        # Create data series
        # prices = Reference(self.ws, min_col=2, min_row=2, max_row=len(data) + 1)
        # dates = Reference(self.ws, min_col=1, min_row=2, max_row=len(data) + 1)

        # Add data to chart
        # chart.add_data(prices, titles_from_data=False)
        # chart.set_categories(dates)

        # Add chart to the sheet
        self.ws.add_chart(chart, "F10")  # Position where the chart will be placed

    def insert_table(self, start_index: int, cols: list, instructions: dict) -> None:
        base_index = start_index + 1

        # Logic to insert header cells.
        header_cells = [f"{c}{start_index}" for c in cols]
        headers = instructions["header"]["values"]
        headers_length = len(headers)
        headers_border = self.create_header_borders(headers_length)

        self.insert_col_range(
            cells=header_cells,
            values=headers,
            fonts=[self.subheader_font] * headers_length,
            fill=self.subheader_fill,
            borders=headers_border,
            alignments=[self.center_align] * headers_length,
        )
        # Logic to insert body cells.
        body = instructions["body"]
        body_font = self.create_body_fonts(headers_length)
        body_borders = self.create_body_borders(headers_length)
        body_alignment = self.create_body_alignment(headers_length)
        for i in body.keys():
            cur_body_cells = [f"{c}{base_index}" for c in cols]
            metric = body[i]
            self.insert_col_range(
                cells=cur_body_cells,
                values=metric["values"],
                fonts=body_font,
                fill=None,
                borders=body_borders,
                alignments=body_alignment,
            )
            base_index += 1

    def insert_col_range(
        self,
        cells: list,
        values: list,
        fonts: list,
        fill: PatternFill,
        borders: list,
        alignments: list,
    ):

        index = 0
        for c in cells:
            self.ws[c].value = values[index]
            self.ws[c].font = fonts[index]
            if fill != None:
                self.ws[c].fill = fill
            if borders[index] != None:
                self.ws[c].border = borders[index]
            self.ws[c].alignment = alignments[index]
            index += 1

    def create_body_alignment(self, length):
        alignments = []
        for i in range(length + 1):
            if i == 0:
                alignments.append(self.left_align)
            else:
                alignments.append(self.right_align)
        return alignments

    def create_body_fonts(self, length: int):
        fonts = []
        for i in range(length + 1):
            if i == 0:
                fonts.append(self.bold_font)
            else:
                fonts.append(self.data_font)
        return fonts

    def create_header_borders(self, length: int):
        borders = []
        for i in range(length + 1):
            if i == 0:
                borders.append(self.top_left_corner_thick_border)
            elif i == length - 1:
                borders.append(self.top_right_corner_thick_border)
            else:
                borders.append(self.top_thick_border)
        return borders

    def create_body_borders(self, length: int):
        borders = []
        for i in range(length + 1):
            if i == 0:
                borders.append(self.left_thick_border)
            elif i == length - 1:
                borders.append(self.right_thick_border)
            else:
                borders.append(None)
        return borders

    def fill_col_borders(
        self, col: str, start_index: int, border: Border, r: int = (6, 18)
    ):
        index = start_index
        for i in range(r[0], r[1] + 1):
            coord = f"{col}{index}"
            self.ws[coord].border = border
            index += 1

    def fill_row_borders(self, index: int, cols: list, border: Border):
        for i in cols:
            coord = f"{i}{index}"
            self.ws[coord].border = border

    def get_x_year_ago(self, year: int = 1):
        today = dt.datetime.now()
        one_year_ago = today.replace(year=today.year - year)
        return one_year_ago


instructions = {
    "header": {
        "values": "",
    },
    "body": {
        "": {
            "values": [],
        },
        "": {
            "values": [],
        },
    },
}
instructions = {
    "header": {
        "values": "",
    },
    "body": {
        "": {
            "values": [],
        },
        "": {
            "values": [],
        },
    },
}
instructions = {
    "header": {
        "values": "",
    },
    "body": {
        "": {
            "values": [],
        },
        "": {
            "values": [],
        },
    },
}
instructions = {
    "header": {
        "values": "",
    },
    "body": {
        "": {
            "values": [],
        },
        "": {
            "values": [],
        },
    },
}
instructions = {
    "header": {
        "values": "",
    },
    "body": {
        "": {
            "values": [],
        },
        "": {
            "values": [],
        },
    },
}
instructions = {
    "header": {
        "values": "",
    },
    "body": {
        "": {
            "values": [],
        },
        "": {
            "values": [],
        },
    },
}
